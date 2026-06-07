"""Tests for GET/POST /api/manifest (Deck Shape editor backend)."""

from __future__ import annotations

import openpyxl
import pytest
from fastapi.testclient import TestClient


def _make_manifest_xlsx(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header = (
        "Slide #", "Slide ID", "Title", "Chart Type",
        "Layout #", "Layout Name", "Slide Type",
        "Keep? (Y/N)", "Your Layout Choice", "Notes",
    )
    ws = wb.create_sheet("ARS - DCTR")
    ws.append(header)
    ws.append((1, "DCTR-1", "Overall DCTR",     "bar", 8, "CUSTOM", "screenshot", "Y", None, None))
    ws.append((2, "DCTR-2", "Open vs Eligible", "bar", 8, "CUSTOM", "screenshot", "",  None, None))
    ws.append((3, "DCTR-X", "Drop me",          "bar", 8, "CUSTOM", "screenshot", "N", None, None))
    ws2 = wb.create_sheet("ARS - RegE")
    ws2.append(header)
    ws2.append((1, "REGE-1", "Reg E Overall", "bar", 8, "CUSTOM", "screenshot", "Y", None, None))
    wb.save(path)


@pytest.fixture
def with_manifest(app_module, tmp_path, monkeypatch):
    path = tmp_path / "SLIDE_MANIFEST.xlsx"
    _make_manifest_xlsx(path)
    monkeypatch.setenv("SLIDE_MANIFEST_PATH", str(path))
    return path


def test_get_manifest_lists_every_row(app_module, with_manifest):
    client = TestClient(app_module.app)
    r = client.get("/api/manifest")
    assert r.status_code == 200
    body = r.json()
    assert body["path"] == str(with_manifest)
    sids = {row["slide_id"] for row in body["rows"]}
    assert sids == {"DCTR-1", "DCTR-2", "DCTR-X", "REGE-1"}
    assert body["counts"]["main"] == 2  # DCTR-1, REGE-1
    assert body["counts"]["drop"] == 1  # DCTR-X
    assert body["counts"]["blank"] == 1  # DCTR-2


def test_post_manifest_persists_updates(app_module, with_manifest):
    client = TestClient(app_module.app)
    r = client.post("/api/manifest", json={"updates": {
        "DCTR-1": "A",   # Y -> A
        "DCTR-2": "Y",   # blank -> Y
        "DCTR-X": "",    # N -> blank
    }})
    assert r.status_code == 200
    body = r.json()
    assert body["updated"] == 3

    # Re-fetch and confirm the new shape
    r2 = client.get("/api/manifest")
    by_id = {row["slide_id"]: row["decision"] for row in r2.json()["rows"]}
    assert by_id["DCTR-1"] == "A"
    assert by_id["DCTR-2"] == "Y"
    assert by_id["DCTR-X"] == ""


def test_post_manifest_rejects_non_dict_payload(app_module, with_manifest):
    client = TestClient(app_module.app)
    r = client.post("/api/manifest", json={"updates": ["not", "a", "dict"]})
    assert r.status_code == 400


def test_post_manifest_no_updates_returns_zero(app_module, with_manifest):
    client = TestClient(app_module.app)
    r = client.post("/api/manifest", json={"updates": {}})
    assert r.status_code == 200
    assert r.json()["updated"] == 0
