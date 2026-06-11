"""Operator-driven slide selection from SLIDE_MANIFEST.xlsx.

The operator maintains a personal `SLIDE_MANIFEST.xlsx` at the M: drive root
(gitignored -- never overwritten by git pull). Each per-section sheet lists
every slide and a `Keep? (Y/N)` column the operator fills in:

    Y       Keep on MAIN deck
    A       Move to Support / Appendix deck (aux routing)
    N       Drop entirely
    blank   Not yet decided -- treated as Y (keep) by default

This module reads the manifest and produces three slide-ID sets the
deck_builder consumes to filter `final_slides`. With a missing or all-blank
manifest the loader returns empty sets, so default behavior is "keep
everything" -- the existing pipeline still runs unchanged until the
operator opts in by marking slides.

Issue: #129. Manifest workflow: SETUP.md, SLIDE_MANIFEST.template.xlsx.
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from pathlib import Path

from loguru import logger


# Sheets that are reference / instructions, not slide rows.
_SKIP_SHEETS = {"Key", "Layout Reference", "Support Deck"}

# Header column positions (1-indexed for openpyxl, 0-indexed for the dict)
_SLIDE_ID_COL = "Slide ID"
_KEEP_COL = "Keep? (Y/N)"


@dataclass(frozen=True)
class ManifestDecisions:
    """Outcome of reading SLIDE_MANIFEST.xlsx."""

    main_ids: frozenset[str] = field(default_factory=frozenset)
    aux_ids: frozenset[str] = field(default_factory=frozenset)
    drop_ids: frozenset[str] = field(default_factory=frozenset)
    undecided_ids: frozenset[str] = field(default_factory=frozenset)
    path_used: str | None = None  # absolute path of the manifest that was read
    error: str | None = None  # set when the file was found but unreadable

    @property
    def is_empty(self) -> bool:
        """True when no Y/A/N decisions are present -- caller should no-op."""
        return not (self.main_ids or self.aux_ids or self.drop_ids)

    def summary(self) -> str:
        """One-line summary suitable for the pipeline log + frontend parsing."""
        if self.error:
            return f"SLIDE MANIFEST: error -- {self.error}"
        if self.path_used is None:
            return "SLIDE MANIFEST: not found (default = keep all slides)"
        return (
            f"SLIDE MANIFEST: main={len(self.main_ids)} "
            f"/ aux={len(self.aux_ids)} "
            f"/ dropped={len(self.drop_ids)} "
            f"/ undecided={len(self.undecided_ids)} "
            f"(from {self.path_used})"
        )


def _candidate_paths() -> list[Path]:
    """Search order for SLIDE_MANIFEST.xlsx, most-preferred first."""
    candidates: list[Path] = []
    # Environment override (tests, ad-hoc runs)
    env = os.environ.get("SLIDE_MANIFEST_PATH")
    if env:
        candidates.append(Path(env))
    # M: drive root (production, Windows)
    if sys.platform == "win32":
        candidates.append(Path(r"M:\ARS\SLIDE_MANIFEST.xlsx"))
    else:
        candidates.append(Path("/Volumes/M/ARS/SLIDE_MANIFEST.xlsx"))
    # Repo root fallback (dev on Mac; operator who copied template)
    here = Path(__file__).resolve()
    for parent in (here.parents[3], here.parents[2]):
        candidates.append(parent / "SLIDE_MANIFEST.xlsx")
    return candidates


def _resolve_manifest_path(path: Path | str | None = None) -> Path | None:
    """Return the first manifest path that exists, or None."""
    if path is not None:
        p = Path(path)
        return p if p.exists() else None
    for cand in _candidate_paths():
        if cand.exists():
            return cand
    return None


def _normalize_decision(raw) -> str | None:
    """Convert a Keep? cell value to one of {Y, A, N} or None for blank."""
    if raw is None:
        return None
    s = str(raw).strip().upper()
    if not s:
        return None
    if s in {"Y", "YES", "KEEP"}:
        return "Y"
    if s in {"A", "AUX", "APPENDIX", "SUPPORT"}:
        return "A"
    if s in {"N", "NO", "DROP", "SKIP"}:
        return "N"
    return None  # ignore unknown values


def load_manifest_decisions(
    path: Path | str | None = None,
) -> ManifestDecisions:
    """Read SLIDE_MANIFEST.xlsx and return the Keep? decisions.

    Returns an empty `ManifestDecisions` (no-op) if the file is missing or
    unreadable. Never raises -- a bad manifest must not break the pipeline.
    """
    resolved = _resolve_manifest_path(path)
    if resolved is None:
        return ManifestDecisions()

    try:
        # Import openpyxl lazily so unit tests that don't touch manifests
        # don't pay the import cost.
        import openpyxl
        wb = openpyxl.load_workbook(str(resolved), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("Could not open SLIDE_MANIFEST.xlsx at {path}: {err}",
                       path=resolved, err=exc)
        return ManifestDecisions(error=str(exc), path_used=str(resolved))

    main: set[str] = set()
    aux: set[str] = set()
    drop: set[str] = set()
    undecided: set[str] = set()

    for sheet_name in wb.sheetnames:
        if sheet_name in _SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        # Find column indices by header label
        try:
            id_idx = header.index(_SLIDE_ID_COL)
            keep_idx = header.index(_KEEP_COL)
        except ValueError:
            # Sheet doesn't have the expected schema -- skip silently.
            continue

        for row in rows:
            if id_idx >= len(row) or keep_idx >= len(row):
                continue
            slide_id = row[id_idx]
            if slide_id is None:
                continue
            sid = str(slide_id).strip()
            if not sid:
                continue
            decision = _normalize_decision(row[keep_idx])
            if decision == "Y":
                main.add(sid)
            elif decision == "A":
                aux.add(sid)
            elif decision == "N":
                drop.add(sid)
            else:
                undecided.add(sid)

    wb.close()
    return ManifestDecisions(
        main_ids=frozenset(main),
        aux_ids=frozenset(aux),
        drop_ids=frozenset(drop),
        undecided_ids=frozenset(undecided),
        path_used=str(resolved),
    )


# ---------------------------------------------------------------------------
# Editor support (Wave 4 follow-up)
#
# read_manifest_rows: list every (sheet, slide_id, title, current_decision)
#   tuple so the UI editor can render a sortable grid.
# write_manifest_decisions: persist a {slide_id: 'Y'|'A'|'N'|''} dict back to
#   SLIDE_MANIFEST.xlsx. Atomic: writes to a temp file, then os.replace.
# ---------------------------------------------------------------------------


_TITLE_COL_CANDIDATES = ("Title", "Title Pattern", "Slide Title", "Headline", "Description")


@dataclass
class ManifestRow:
    sheet: str = ""
    slide_id: str = ""
    title: str = ""
    decision: str = ""  # "Y" | "A" | "N" | ""


def read_manifest_rows(path: Path | str | None = None) -> list[ManifestRow]:
    """Return every slide row across every per-section sheet in SLIDE_MANIFEST.xlsx."""
    resolved = _resolve_manifest_path(path)
    if resolved is None:
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(resolved), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("read_manifest_rows: cannot open {p}: {err}", p=resolved, err=exc)
        return []

    out = _rows_from_workbook(wb)
    wb.close()
    return out


# Sheet routing: slide_id -> per-section sheet name, matching
# SLIDE_MANIFEST.template.xlsx. TXN codes per analytics.txn_wrapper.TXN_SECTIONS.
_TXN_CODE_SHEETS = {
    "GEN": "TXN - General",
    "MERCH": "TXN - Merchant",
    "MCC": "TXN - MCC Code",
    "BUS": "TXN - Business",
    "PERS": "TXN - Personal",
    "COMP": "TXN - Competition",
    "FIN": "TXN - Financial Svc",
    "ICSA": "TXN - ICS Acquisition",
    "CAMP": "TXN - Campaign",
    "BR": "TXN - Branch",
    "TT": "TXN - TXN Type",
    "PROD": "TXN - Product",
    "ATR": "TXN - Attrition",
    "BAL": "TXN - Balance",
    "IC": "TXN - Interchange",
    "REGE": "TXN - RegE Overdraft",
    "PAY": "TXN - Payroll",
    "REL": "TXN - Relationship",
    "SEG": "TXN - Segment Evol",
    "RET": "TXN - Retention",
    "ENG": "TXN - Engagement",
    "EXEC": "TXN - Executive",
}


def sheet_for_slide(slide_id: str, module_id: str = "") -> str:
    """Template sheet name a slide row belongs on, from its slide_id prefix."""
    sid = (slide_id or "").upper()
    if sid.startswith("TXN-"):
        code = sid.split("-")[1] if "-" in sid[4:] else sid[4:]
        return _TXN_CODE_SHEETS.get(code, f"TXN - {code.title()}")
    if sid.startswith("DCTR") or sid.startswith("A7"):
        return "ARS - DCTR"
    if sid.startswith("REGE") or sid.startswith("A8"):
        return "ARS - RegE"
    if sid.startswith("A9") or "ATTRITION" in sid:
        return "ARS - Attrition"
    if sid.startswith("A10") or sid.startswith("A11") or "VALUE" in sid:
        return "ARS - Value"
    if sid[:3] in {"A12", "A13", "A14", "A15", "A16", "A17"} or "MAILER" in sid:
        return "ARS - Mailer"
    if sid.startswith("A18") or sid.startswith("A19") or sid.startswith("A20") \
            or sid.startswith("S") or "INSIGHT" in sid:
        return "ARS - Insights"
    if sid.startswith("ICS"):
        return "ARS - ICS"
    if sid.startswith("A1") or "OVERVIEW" in sid:
        return "ARS - Overview"
    return "Other"


_MANIFEST_HEADER = (
    "Slide #", "Slide ID", "Title", "Chart Type", "Layout #",
    "Layout Name", "Slide Type", "Keep? (Y/N)", "Your Layout Choice", "Notes",
)


def _writable_manifest_target() -> Path | None:
    """Where to create SLIDE_MANIFEST.xlsx when none exists yet."""
    for cand in _candidate_paths():
        if cand.parent.exists():
            return cand
    return None


def ensure_manifest_rows(
    slides: list[dict],
    path: Path | str | None = None,
) -> tuple[str | None, int]:
    """Append rows for slides not yet present in SLIDE_MANIFEST.xlsx.

    `slides`: dicts with at least slide_id, optionally title and module_id.
    Slides already listed on any sheet are left untouched (the operator's
    Keep? decisions survive). Creates the workbook -- and any missing
    per-section sheet -- on demand, so a fresh run can be synced without
    copying the template first. Returns (manifest_path, rows_added).
    """
    import openpyxl

    resolved = _resolve_manifest_path(path)
    creating = resolved is None
    if creating:
        resolved = Path(path) if path is not None else _writable_manifest_target()
        if resolved is None:
            logger.warning("ensure_manifest_rows: no writable location for SLIDE_MANIFEST.xlsx")
            return None, 0

    if creating:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    else:
        try:
            wb = openpyxl.load_workbook(str(resolved), read_only=False, data_only=False)
        except Exception as exc:
            logger.warning("ensure_manifest_rows: cannot open {p}: {err}", p=resolved, err=exc)
            return str(resolved), 0

    existing = {r.slide_id for r in _rows_from_workbook(wb)}

    added = 0
    for slide in slides:
        sid = str(slide.get("slide_id", "") or "").strip()
        if not sid or sid in existing:
            continue
        sheet_name = sheet_for_slide(sid, str(slide.get("module_id", "") or ""))
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(_MANIFEST_HEADER)
        ws.append((
            ws.max_row,  # header is row 1, so this numbers slides from 1
            sid,
            str(slide.get("title", "") or ""),
            None, None, None,
            str(slide.get("slide_type", "") or "") or None,
            None,  # Keep? -- blank = undecided (default keep)
            None, None,
        ))
        existing.add(sid)
        added += 1

    if added == 0 and not creating:
        wb.close()
        return str(resolved), 0

    import tempfile
    target = Path(resolved)
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, dir=str(target.parent)
    ) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(str(tmp_path))
        os.replace(str(tmp_path), str(target))
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
    wb.close()
    logger.info("ensure_manifest_rows: {n} row(s) added to {p}", n=added, p=resolved)
    return str(target), added


def _rows_from_workbook(wb) -> list[ManifestRow]:
    """ManifestRow list from an already-open workbook (any mode)."""
    out: list[ManifestRow] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in _SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        try:
            id_idx = header.index(_SLIDE_ID_COL)
            keep_idx = header.index(_KEEP_COL)
        except ValueError:
            continue
        title_idx: int | None = None
        for cand in _TITLE_COL_CANDIDATES:
            if cand in header:
                title_idx = header.index(cand)
                break
        for row in rows:
            if id_idx >= len(row) or keep_idx >= len(row):
                continue
            slide_id = row[id_idx]
            if slide_id is None:
                continue
            sid = str(slide_id).strip()
            if not sid:
                continue
            title_val = ""
            if title_idx is not None and title_idx < len(row) and row[title_idx] is not None:
                title_val = str(row[title_idx]).strip()
            out.append(ManifestRow(
                sheet=sheet_name,
                slide_id=sid,
                title=title_val,
                decision=_normalize_decision(row[keep_idx]) or "",
            ))
    return out


def write_manifest_decisions(
    updates: dict[str, str],
    path: Path | str | None = None,
) -> int:
    """Write Keep? decisions back to SLIDE_MANIFEST.xlsx. Returns rows updated.

    `updates` maps slide_id -> "Y" | "A" | "N" | "" (blank clears the decision).
    Unknown decision values are ignored. Atomic write via temp + os.replace.
    No-op if the manifest can't be located.
    """
    if not updates:
        return 0
    resolved = _resolve_manifest_path(path)
    if resolved is None:
        logger.warning("write_manifest_decisions: SLIDE_MANIFEST.xlsx not found")
        return 0

    # Normalize decisions to {Y, A, N, ""}
    normalized: dict[str, str] = {}
    for sid, val in updates.items():
        if not sid:
            continue
        if val == "" or val is None:
            normalized[str(sid).strip()] = ""
            continue
        canonical = _normalize_decision(val)
        normalized[str(sid).strip()] = canonical or ""

    import openpyxl
    import os
    import tempfile

    try:
        wb = openpyxl.load_workbook(str(resolved), read_only=False, data_only=False)
    except Exception as exc:
        logger.warning("write_manifest_decisions: cannot open {p}: {err}", p=resolved, err=exc)
        return 0

    updated = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in _SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            continue
        try:
            id_idx = header_row.index(_SLIDE_ID_COL) + 1  # openpyxl is 1-based
            keep_idx = header_row.index(_KEEP_COL) + 1
        except ValueError:
            continue
        for row_idx in range(2, ws.max_row + 1):
            sid_cell = ws.cell(row=row_idx, column=id_idx).value
            if sid_cell is None:
                continue
            sid = str(sid_cell).strip()
            if sid in normalized:
                ws.cell(row=row_idx, column=keep_idx).value = normalized[sid] or None
                updated += 1

    # Atomic write
    target = Path(resolved)
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, dir=str(target.parent)
    ) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(str(tmp_path))
        os.replace(str(tmp_path), str(target))
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
    wb.close()
    logger.info("write_manifest_decisions: {n} row(s) updated in {p}", n=updated, p=resolved)
    return updated
