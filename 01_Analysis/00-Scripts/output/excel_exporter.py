"""CSM review summary workbook (Phase T3.1 / issue #156).

Emits ``[client]_[month]_review_summary.xlsx`` with four sheets the CSM
uses to verify numbers before approving the deck. Source of truth for
what goes in each sheet is the PRD attached to #145 and ``docs/deck/
ARS_SLIDE_DESIGN_CHECKLIST.md``.

Sheets
======
1. **Slide Inventory** — every SlideContent + every dropped record.
2. **KPI Summary** — the headline metrics the CSM verifies first.
3. **Callout Text** — populated callout strings for spot-checking.
4. **Data Quality Flags** — drop reasons, soft failures, manifest hits.

Failure mode
============
Workbook write failure is logged at WARNING and never raises; the
pipeline keeps shipping the PPTX even if this export breaks. The
quality gate (T3.2) records the absence in its report.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Styling constants -- aligned to SLIDE_DESIGN.md §4 / §5 so the workbook
# reads as a sibling artifact to the deck.
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="1E3D59")    # navy
_HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
_FLAG_FILL = PatternFill("solid", fgColor="FDECEA")      # light red for flags


@dataclass
class ExcelExportResult:
    path: Path
    sheets_written: list[str]
    rows_written: dict[str, int]


# ---------------------------------------------------------------------------
# Main entry
# ---------------------------------------------------------------------------


class ExcelExporter:
    """Emit the 4-sheet review workbook for a single client run."""

    @classmethod
    def export(cls, ctx, out_path: Path) -> ExcelExportResult | None:
        """Build and save the workbook. Returns None on failure."""
        try:
            wb = Workbook()
            # openpyxl creates a default empty sheet; rename + use it as sheet 1.
            sh1 = wb.active
            sh1.title = "Slide Inventory"
            cls._sheet_inventory(sh1, ctx)

            sh2 = wb.create_sheet("KPI Summary")
            cls._sheet_kpis(sh2, ctx)

            sh3 = wb.create_sheet("Callout Text")
            cls._sheet_callouts(sh3, ctx)

            sh4 = wb.create_sheet("Data Quality Flags")
            cls._sheet_quality_flags(sh4, ctx)

            out_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(out_path)
            return ExcelExportResult(
                path=out_path,
                sheets_written=["Slide Inventory", "KPI Summary",
                                "Callout Text", "Data Quality Flags"],
                rows_written={
                    "Slide Inventory": sh1.max_row - 1,
                    "KPI Summary": sh2.max_row - 1,
                    "Callout Text": sh3.max_row - 1,
                    "Data Quality Flags": sh4.max_row - 1,
                },
            )
        except (OSError, ValueError, KeyError) as exc:
            logger.warning("ExcelExporter failed: {err}", err=exc)
            print(f"  ExcelExporter failed: {exc}")
            return None

    # -----------------------------------------------------------------
    # Sheet builders
    # -----------------------------------------------------------------

    @classmethod
    def _sheet_inventory(cls, sh, ctx) -> None:
        headers = ["Slide ID", "Section", "Type", "Title",
                   "Chart?", "Excel data?", "Status", "Drop reason"]
        cls._write_header(sh, headers)

        # Live slides
        all_slides = getattr(ctx, "all_slides", []) or []
        for slide in all_slides:
            sid = getattr(slide, "slide_id", "") or ""
            title = (getattr(slide, "title", "") or "")[:200]
            section = _section_from_slide_id(sid)
            slide_type = _safe_str(getattr(slide, "slide_type", "")) or "screenshot"
            has_chart = "yes" if (getattr(slide, "chart_path", None) and Path(str(slide.chart_path)).exists()) else "no"
            has_excel = "yes" if getattr(slide, "excel_data", None) else "no"
            status = "ok" if getattr(slide, "success", True) else "failed"
            sh.append([sid, section, slide_type, title, has_chart, has_excel, status, ""])

        # Dropped records
        for drop in getattr(ctx, "dropped_slides", None) or []:
            sh.append([
                drop.get("slide_id", ""),
                _section_from_slide_id(drop.get("slide_id", "")),
                "(dropped)", "", "no", "no", "dropped",
                f"{drop.get('reason', '')}: {drop.get('detail', '')}".strip(": "),
            ])

        cls._autofit(sh, [12, 12, 14, 60, 8, 12, 10, 36])

    @classmethod
    def _sheet_kpis(cls, sh, ctx) -> None:
        headers = ["Metric", "Value", "Period", "Comparison", "vs Peer"]
        cls._write_header(sh, headers)
        results = getattr(ctx, "results", {}) or {}
        client = getattr(ctx, "client", None)
        period = getattr(client, "month", "") if client else ""

        rows = [
            ("DCTR — current rate", _pct(_get(results, "dctr_1.rate"))),
            ("DCTR — eligible accounts", _intc(_get(results, "dctr_1.eligible_count"))),
            ("Reg E — opt-in rate", _pct(_get(results, "reg_e_1.rate"))),
            ("Reg E — opt-ins (count)", _intc(_get(results, "reg_e_1.opt_in_count"))),
            ("Attrition — overall rate", _pct(_get(results, "attrition_1.overall_rate"))),
            ("Attrition — closures (count)", _intc(_get(results, "attrition_1.n_closed"))),
            ("Value — total opportunity", _usd_m(_get(results, "value_summary.total"))),
            ("Value — DCTR gap opportunity", _usd_m(_get(results, "value_1.dctr_gap_value"))),
            ("Value — Reg E gap opportunity", _usd_m(_get(results, "value_2.opportunity"))),
            ("Mailer — campaigns in window", _intc(_get(results, "mailer_summary.n_campaigns"))),
            ("Mailer — response rate", _pct(_get(results, "mailer_summary.response_rate"))),
            ("Mailer — ROI multiple", _multiple(_get(results, "mailer_roi.roi_multiple"))),
        ]
        for metric, value in rows:
            comparison = ""
            vs_peer = ""
            if metric.startswith("DCTR"):
                vs_peer = _pp(_get(results, "dctr_1.peer_gap_pp"))
            elif metric.startswith("Reg E"):
                vs_peer = _pp(_get(results, "reg_e_1.peer_gap_pp"))
            elif metric.startswith("Attrition"):
                vs_peer = _pp(_get(results, "attrition_peer.gap_pp"))
            sh.append([metric, value, period, comparison, vs_peer])

        cls._autofit(sh, [38, 14, 12, 24, 14])

    @classmethod
    def _sheet_callouts(cls, sh, ctx) -> None:
        headers = ["Slide ID", "Section", "Metric", "Value",
                   "Denominator", "Comparison", "Insight"]
        cls._write_header(sh, headers)

        for slide in getattr(ctx, "all_slides", []) or []:
            callout = getattr(slide, "callout_box", None)
            if callout is None:
                continue
            sh.append([
                getattr(slide, "slide_id", "") or "",
                getattr(slide, "section_key", "") or "",
                getattr(callout, "metric", ""),
                getattr(callout, "value", ""),
                getattr(callout, "denominator", ""),
                getattr(callout, "comparison", ""),
                getattr(callout, "insight", ""),
            ])

        # Fall back to KPI-derived callouts when the explicit dataclass
        # isn't set (legacy slides). Mirrors CalloutBoxBuilder.from_kpis.
        for slide in getattr(ctx, "all_slides", []) or []:
            if getattr(slide, "callout_box", None) is not None:
                continue
            kpis = getattr(slide, "kpis", None) or {}
            hero_label = hero_value = None
            for label, value in kpis.items():
                if str(label).lower() in ("subtitle", "title"):
                    continue
                if hero_value is None:
                    hero_label, hero_value = str(label), str(value)
                    break
            if not hero_value:
                continue
            sh.append([
                getattr(slide, "slide_id", "") or "",
                getattr(slide, "section_key", "") or "",
                hero_label or "",
                hero_value,
                "(derived from kpis dict)", "", "",
            ])

        cls._autofit(sh, [10, 12, 28, 14, 26, 24, 36])

    @classmethod
    def _sheet_quality_flags(cls, sh, ctx) -> None:
        headers = ["Severity", "Flag", "Slide / Section", "Reason / Detail", "Action"]
        cls._write_header(sh, headers)

        for drop in getattr(ctx, "dropped_slides", None) or []:
            sid = drop.get("slide_id", "")
            reason = drop.get("reason", "")
            severity = _severity_for_reason(reason)
            sh.append([
                severity,
                f"Drop: {reason}",
                sid,
                drop.get("detail", "") or "",
                _action_for_reason(reason),
            ])
            if severity == "high":
                for cell in sh[sh.max_row]:
                    cell.fill = _FLAG_FILL

        # Failed AnalysisResult instances that survived to deck-build
        # (build_deck emits SOFT FAILURE lines for these; surface them
        # in the workbook too so the CSM sees the slide-level errors
        # in one place).
        for r in getattr(ctx, "all_slides", []) or []:
            if not getattr(r, "success", True):
                sh.append([
                    "medium",
                    "Module failed (kept slide for diagnostics)",
                    getattr(r, "slide_id", "") or "",
                    (getattr(r, "error", "") or "")[:240],
                    "Investigate analytics module; rerun if data is now available.",
                ])

        if sh.max_row == 1:
            sh.append(["—", "No data quality flags raised", "", "", ""])

        cls._autofit(sh, [10, 30, 18, 60, 40])

    # -----------------------------------------------------------------
    # Helpers
    # -----------------------------------------------------------------

    @staticmethod
    def _write_header(sh, headers: list[str]) -> None:
        sh.append(headers)
        for cell in sh[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
        sh.freeze_panes = "A2"

    @staticmethod
    def _autofit(sh, widths: list[int]) -> None:
        for idx, w in enumerate(widths, start=1):
            sh.column_dimensions[get_column_letter(idx)].width = w
        # Body cells: left/top + wrap.
        for row in sh.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = _BODY_ALIGN


# ---------------------------------------------------------------------------
# Lookups + formatters
# ---------------------------------------------------------------------------


def _get(results: dict, dotted: str, default: Any = None) -> Any:
    obj: Any = results
    for seg in dotted.split("."):
        if isinstance(obj, dict):
            obj = obj.get(seg)
        else:
            obj = getattr(obj, seg, None)
        if obj is None:
            return default
    return obj


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v)


def _pct(v: Any) -> str:
    if not _is_num(v):
        return ""
    return f"{float(v) * 100:.1f}%"


def _pp(v: Any) -> str:
    if not _is_num(v):
        return ""
    sign = "−" if float(v) < 0 else "+"
    return f"{sign}{abs(float(v)) * 100:.1f} pp"


def _intc(v: Any) -> str:
    if not _is_num(v):
        return ""
    return f"{int(v):,}"


def _usd_m(v: Any) -> str:
    if not _is_num(v):
        return ""
    val = float(v)
    return f"${abs(val) / 1_000_000:.2f}M"


def _multiple(v: Any) -> str:
    if not _is_num(v):
        return ""
    return f"{float(v):.1f}x"


def _is_num(v: Any) -> bool:
    if not isinstance(v, (int, float)):
        return False
    try:
        import math
        return not (math.isnan(v) or math.isinf(v))
    except (TypeError, ValueError):
        return False


def _section_from_slide_id(sid: str) -> str:
    if not sid:
        return ""
    if sid.startswith("_section:"):
        return sid.split(":", 1)[1]
    prefix = sid.split("-")[0].split(".")[0].lower()
    return prefix


def _severity_for_reason(reason: str) -> str:
    if reason in ("module_failed",):
        return "high"
    if reason in ("data_missing",):
        return "medium"
    return "low"


def _action_for_reason(reason: str) -> str:
    return {
        "data_missing": "Verify the analytics module ran and emitted results for this slide.",
        "client_no_product": "Confirm client doesn't have this product. Update 03_Config if they do.",
        "threshold_not_met": "Module gated output on a minimum count; expected when client is small.",
        "module_failed": "Investigate analytics module; rerun if data is now available.",
        "manifest_dropped": "SLIDE_MANIFEST.xlsx marked this slide N; restore by setting it to Y.",
        "section_inactive": "Operator excluded this section in the UI; rerun without the exclusion.",
    }.get(reason, "")
