"""Step: Generate deliverables (Excel, PowerPoint, archive)."""

from __future__ import annotations

import json
import shutil
from dataclasses import asdict, dataclass

import openpyxl
from loguru import logger

from ars_analysis.output.excel_formatter import (
    create_summary_sheet,
    format_worksheet,
)
from ars_analysis.pipeline.context import PipelineContext


@dataclass
class SlideStatus:
    """Diagnostic status for a single slide in the run report."""

    slide_id: str
    module_id: str
    success: bool
    has_chart: bool
    has_excel: bool
    error: str = ""
    title: str = ""
    # Wave 4 follow-up: persist the chart_path so a deck-only rebuild
    # (or any post-hoc tooling) can find the PNG without re-running
    # the analytics module. Empty when has_chart is False.
    chart_path: str = ""
    layout_index: int = 8
    slide_type: str = "screenshot"


def _txn_suffix(ctx: PipelineContext) -> str:
    """'_txn' for TXN runs, '' otherwise -- keeps ARS artifact names legacy
    while letting ARS + TXN run concurrently in one client folder."""
    product = (getattr(ctx, "product", None)
               or (getattr(ctx.settings, "product", "") if ctx.settings else "")
               or "")
    return "_txn" if str(product).lower() == "txn" else ""


def _build_run_report(ctx: PipelineContext) -> list[SlideStatus]:
    """Build a diagnostic report of all slide statuses after analysis."""
    report: list[SlideStatus] = []
    for result in ctx.all_slides:
        module_id = ""
        # Determine module_id from ctx.results mapping
        for mid, results_list in ctx.results.items():
            if isinstance(results_list, list) and result in results_list:
                module_id = mid
                break

        chart_path_str = ""
        if result.chart_path and result.chart_path.exists():
            chart_path_str = str(result.chart_path)
        report.append(
            SlideStatus(
                slide_id=result.slide_id,
                module_id=module_id,
                success=result.success,
                has_chart=bool(result.chart_path and result.chart_path.exists()),
                has_excel=bool(result.excel_data),
                error=result.error or "",
                title=result.title,
                chart_path=chart_path_str,
                layout_index=getattr(result, "layout_index", 8),
                slide_type=getattr(result, "slide_type", "screenshot"),
            )
        )
    return report


def rebuild_deck_from_report(ctx: PipelineContext) -> None:
    """Rebuild the PPTX from a previously persisted run_report.json.

    Skips analyze + Excel writes; reconstructs minimal AnalysisResult stubs
    from the saved (slide_id, chart_path, title) tuples and hands them to
    build_deck. Used by POST /api/rebuild_deck after the operator edits
    SLIDE_MANIFEST.xlsx via the in-UI Deck Shape editor.

    Returns silently; deck path is added to ctx.export_log on success.
    """
    report_path = ctx.paths.base_dir / (
        f"{ctx.client.client_id}_{ctx.client.month}{_txn_suffix(ctx)}_run_report.json"
    )
    if not report_path.exists():
        # Fall back to the most recent report of either product
        _cands = sorted(
            ctx.paths.base_dir.glob(
                f"{ctx.client.client_id}_{ctx.client.month}*_run_report.json"
            ),
            key=lambda f: f.stat().st_mtime,
            reverse=True,
        )
        if _cands:
            report_path = _cands[0]
    if not report_path.exists():
        logger.warning("rebuild_deck_from_report: {p} not found", p=report_path)
        return
    payload = json.loads(report_path.read_text(encoding="utf-8"))
    slides = payload.get("slides") or []

    from ars_analysis.analytics.base import AnalysisResult
    from pathlib import Path as _Path

    stubs: list[AnalysisResult] = []
    for s in slides:
        if not s.get("has_chart"):
            continue
        chart = s.get("chart_path") or ""
        if not chart or not _Path(chart).exists():
            continue
        stubs.append(AnalysisResult(
            slide_id=s.get("slide_id", ""),
            title=s.get("title", ""),
            chart_path=_Path(chart),
            success=True,
            layout_index=int(s.get("layout_index") or 8),
            slide_type=s.get("slide_type") or "screenshot",
        ))

    ctx.all_slides = stubs
    # Replay the slide_id keyed results so spec rendering still works.
    ctx.results = ctx.results or {}
    logger.info("rebuild_deck_from_report: loaded {n} slides from {p}", n=len(stubs), p=report_path)
    if not stubs:
        return
    _build_deck(ctx)
    _build_aux_deck(ctx)


def _save_run_report(ctx: PipelineContext, report: list[SlideStatus]) -> None:
    """Save run report as JSON next to output files."""
    report_path = ctx.paths.base_dir / (
        f"{ctx.client.client_id}_{ctx.client.month}{_txn_suffix(ctx)}_run_report.json"
    )
    ctx.paths.base_dir.mkdir(parents=True, exist_ok=True)

    ok = sum(1 for s in report if s.success and s.has_chart)
    failed = sum(1 for s in report if not s.success)
    no_chart = sum(1 for s in report if s.success and not s.has_chart)

    report_data = {
        "client_id": ctx.client.client_id,
        "month": ctx.client.month,
        "summary": {
            "total": len(report),
            "ok": ok,
            "failed": failed,
            "no_chart": no_chart,
        },
        "slides": [asdict(s) for s in report],
    }
    report_path.write_text(json.dumps(report_data, indent=2))
    ctx.export_log.append(str(report_path))
    logger.info(
        "Run report: {ok}/{total} slides OK, {failed} failed, {no_chart} no chart",
        ok=ok,
        total=len(report),
        failed=failed,
        no_chart=no_chart,
    )


def _drop_empty_slides(ctx: PipelineContext) -> int:
    """G6: silently skip slides whose underlying data is empty.

    Empty = ran successfully but produced neither a chart nor Excel data.
    Mutates ctx.all_slides in place. Returns the number dropped.
    """
    before = len(ctx.all_slides)
    kept = []
    dropped_ids: list[str] = []
    for r in ctx.all_slides:
        is_empty = (
            getattr(r, "success", True)
            and not (getattr(r, "chart_path", None) and r.chart_path.exists())
            and not getattr(r, "excel_data", None)
        )
        if is_empty:
            dropped_ids.append(getattr(r, "slide_id", "?"))
            continue
        kept.append(r)
    ctx.all_slides = kept
    dropped = before - len(kept)
    if dropped:
        logger.info(
            "G6 drop-if-empty: skipped {n} slide(s) with empty data: {ids}",
            n=dropped,
            ids=", ".join(dropped_ids[:20]) + ("..." if len(dropped_ids) > 20 else ""),
        )
    return dropped


def step_generate(ctx: PipelineContext) -> None:
    """Generate all output deliverables from analysis results.

    Order: drop-if-empty -> run report -> Excel workbook -> PowerPoint deck (main + aux).
    Uses single-write pattern: build Excel once, then shutil.copy2 for master.
    """
    if not ctx.all_slides:
        logger.warning("No analysis results to generate deliverables from")
        return

    # G6: drop slides with empty data before any deliverable is produced.
    _drop_empty_slides(ctx)

    if not ctx.all_slides:
        logger.warning("No slides remain after G6 drop-if-empty filter")
        return

    # Build and save run report before deck build (diagnostics first)
    report = _build_run_report(ctx)
    _save_run_report(ctx, report)
    ctx.results["_run_report"] = report

    _write_excel(ctx)
    _build_deck(ctx)
    _build_aux_deck(ctx)
    logger.info("Deliverables generated for {client}", client=ctx.client.client_id)

    # Wave 1: denominator-law audit. Walks ctx.all_slides, writes rates_audit.csv,
    # flags any rate whose denominator label is not in the 4-layer framework.
    # Runs before the scorecard so violations are visible in the scorecard summary.
    try:
        from ars_analysis.pipeline.steps.audit import step_audit
        step_audit(ctx)
    except Exception as _exc:
        logger.warning("rates_audit step failed: {err}", err=_exc)

    # Post-run scorecard derived from the manifest. Optional: only runs if
    # the manifest was set up. Failures here never break the pipeline.
    _mf = getattr(ctx, "manifest", None)
    if _mf is not None:
        try:
            from ars_analysis.pipeline.scorecard import write as _scorecard_write
            _scorecard_path = ctx.paths.base_dir / f"run_scorecard{_txn_suffix(ctx)}.md"
            _scorecard_write(_mf, _scorecard_path)
            logger.info("Run scorecard written: {p}", p=_scorecard_path)
        except Exception as _exc:
            logger.warning("scorecard write failed: {err}", err=_exc)


def step_archive(ctx: PipelineContext) -> None:
    """Copy deliverables to archive location. Non-critical step."""
    logger.info("Archive step for {client} (not yet implemented)", client=ctx.client.client_id)


def _write_excel(ctx: PipelineContext) -> None:
    """Write all analysis results to a formatted Excel workbook.

    Single-write pattern: one workbook with a tab per analysis.
    Then shutil.copy2 for the master/archive copy.
    """
    excel_path = ctx.paths.excel_dir / (
        f"{ctx.client.client_id}_{ctx.client.month}{_txn_suffix(ctx)}_analysis.xlsx"
    )
    ctx.paths.excel_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheets_written = 0
    for result in ctx.all_slides:
        if result.excel_data is None:
            continue
        for sheet_name, df in result.excel_data.items():
            # Truncate sheet name to Excel 31-char limit
            safe_name = f"{result.slide_id}_{sheet_name}"[:31]
            ws = wb.create_sheet(title=safe_name)
            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            # Write data rows
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            # Format the worksheet
            format_worksheet(ws)
            sheets_written += 1

    if sheets_written == 0:
        logger.warning("No Excel data to write")
        return

    # Add summary sheet at the front
    create_summary_sheet(wb, ctx)

    wb.save(excel_path)
    ctx.export_log.append(str(excel_path))
    logger.info("Excel written: {path} ({n} sheets)", path=excel_path.name, n=sheets_written)

    # Single-write pattern: copy to master location if configured
    if ctx.settings and hasattr(ctx.settings, "paths"):
        master_dir = getattr(ctx.settings.paths, "ars_base", None)
        if master_dir and master_dir != ctx.paths.excel_dir:
            master_path = master_dir / excel_path.name
            try:
                shutil.copy2(excel_path, master_path)
                logger.info("Master copy: {name}", name=master_path.name)
            except OSError as exc:
                logger.warning("Master copy failed: {err}", err=exc)


def _build_deck(ctx: PipelineContext) -> None:
    """Build PowerPoint deck from analysis results."""
    skip_pptx = False
    if ctx.settings and hasattr(ctx.settings, "pipeline"):
        skip_pptx = getattr(ctx.settings.pipeline, "skip_pptx", False)

    if skip_pptx:
        logger.info("PowerPoint generation skipped (skip_pptx=True)")
        return

    try:
        from ars_analysis.output.deck_builder import build_deck

        result = build_deck(ctx)
        if result:
            if ctx.progress_callback:
                ctx.progress_callback(f"PPTX deck: {result.name}")
        else:
            msg = "PPTX: no slides with charts to build deck"
            logger.warning(msg)
            if ctx.progress_callback:
                ctx.progress_callback(msg)
    except ImportError:
        msg = f"PPTX skipped: deck_builder not available ({len(ctx.all_slides)} slides ready)"
        logger.warning(msg)
        if ctx.progress_callback:
            ctx.progress_callback(msg)
    except Exception as exc:
        msg = f"PPTX build failed: {exc}"
        logger.error(msg)
        ctx.export_log.append(f"ERROR: {msg}")
        if ctx.progress_callback:
            ctx.progress_callback(msg)


def _build_aux_deck(ctx: PipelineContext) -> None:
    """G7: build a secondary auxiliary deck containing slides routed away from main.

    Routing source: ctx.auxiliary_slide_ids (set of slide_ids).
    No-op when the set is empty (Wave 1 ships empty; later waves populate).
    """
    aux_ids = getattr(ctx, "auxiliary_slide_ids", set()) or set()
    if not aux_ids:
        logger.debug("G7 aux deck: no slide IDs routed to aux; skipping")
        return

    skip_pptx = False
    if ctx.settings and hasattr(ctx.settings, "pipeline"):
        skip_pptx = getattr(ctx.settings.pipeline, "skip_pptx", False)
    if skip_pptx:
        return

    aux_slides = [r for r in ctx.all_slides if getattr(r, "slide_id", "") in aux_ids]
    if not aux_slides:
        logger.info("G7 aux deck: routing set non-empty but no matching slides; skipping")
        return

    # Swap ctx.all_slides for the aux build, then restore.
    main_slides = ctx.all_slides
    try:
        ctx.all_slides = aux_slides
        from ars_analysis.output.deck_builder import build_deck

        # Mark this build as aux via a transient attribute the builder can read
        # to choose the output filename. Fallback: rename here if the builder
        # does not honor the flag.
        ctx._aux_build = True  # type: ignore[attr-defined]
        result = build_deck(ctx)
        if result and not result.name.endswith("_aux_deck.pptx"):
            # Builder did not honor _aux_build; rename in-place.
            aux_path = result.with_name(result.name.replace("_deck.pptx", "_aux_deck.pptx"))
            try:
                result.rename(aux_path)
                if str(result) in ctx.export_log:
                    ctx.export_log[ctx.export_log.index(str(result))] = str(aux_path)
                result = aux_path
            except OSError as exc:
                logger.warning("Aux deck rename failed: {err}", err=exc)
        if result:
            logger.info("G7 aux deck built: {name} ({n} slides)", name=result.name, n=len(aux_slides))
            if ctx.progress_callback:
                ctx.progress_callback(f"Aux deck: {result.name}")
    except Exception as exc:
        logger.error("G7 aux deck build failed: {err}", err=exc)
    finally:
        ctx.all_slides = main_slides
        if hasattr(ctx, "_aux_build"):
            delattr(ctx, "_aux_build")
